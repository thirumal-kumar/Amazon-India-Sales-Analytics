import io
import streamlit as st
import pandas as pd
from utils import load_data, filter_controls, page_title, kpi_card

# Optional: PowerPoint support (not essential for Excel fix)
try:
    from pptx import Presentation
    from pptx.util import Inches
    PPTX_AVAILABLE = True
except:
    PPTX_AVAILABLE = False

# -----------------------------
# PAGE TITLE
# -----------------------------
page_title("📤 Report Export (Excel & PPTX)", "Download KPIs, summaries, and key tables")

# -----------------------------
# LOAD DATA
# -----------------------------
df = load_data()
if df is None or df.empty:
    st.warning("⚠ Please upload data from the Home page.")
    st.stop()

# Convert date column
if "order_date" in df.columns:
    df["order_date"] = pd.to_datetime(df["order_date"], errors="coerce")

df_f = filter_controls(df)
if df_f.empty:
    st.info("ℹ No data available after applying filters.")
    st.stop()

# Identify revenue column
rev_col = None
for c in ["final_amount_inr", "order_amount", "subtotal_inr", "revenue"]:
    if c in df_f.columns:
        rev_col = c
        break

if rev_col is None:
    st.error("❌ No valid revenue column found (like 'final_amount_inr').")
    st.stop()

df_f[rev_col] = pd.to_numeric(df_f[rev_col], errors="coerce").fillna(0)

# -----------------------------
# KPI CALCULATIONS
# -----------------------------
total_revenue = df_f[rev_col].sum()
total_orders = len(df_f)
unique_customers = df_f["customer_id"].nunique() if "customer_id" in df_f else 0
aov = total_revenue / total_orders if total_orders else 0

c1, c2, c3, c4 = st.columns(4)
kpi_card("Total Revenue", f"₹{total_revenue:,.0f}", column=c1)
kpi_card("Total Orders", f"{total_orders:,}", column=c2)
kpi_card("Unique Customers", f"{unique_customers:,}", column=c3)
kpi_card("Avg Order Value", f"₹{aov:,.0f}", column=c4)

# -----------------------------
# SUMMARY TABLES (OPTIONAL SHEETS)
# -----------------------------
# Monthly revenue
if "order_date" in df_f.columns:
    monthly = df_f.copy()
    monthly["Month"] = monthly["order_date"].dt.to_period("M").astype(str)
    monthly = monthly.groupby("Month")[rev_col].sum().reset_index()
else:
    monthly = pd.DataFrame(columns=["Month", "Revenue"])

# Category revenue
if "category" in df_f.columns:
    cat_df = df_f.groupby("category")[rev_col].sum().reset_index().rename(columns={rev_col: "Revenue"})
else:
    cat_df = pd.DataFrame(columns=["category", "Revenue"])

# Product revenue
prod_col = next((c for c in ["product_name", "product", "product_id", "title"] if c in df_f.columns), None)
if prod_col:
    prod_df = df_f.groupby(prod_col)[rev_col].sum().reset_index().rename(columns={prod_col: "Product", rev_col: "Revenue"})
else:
    prod_df = pd.DataFrame(columns=["Product", "Revenue"])

# State revenue
state_col = next((c for c in ["customer_state", "state", "ship_state"] if c in df_f.columns), None)
if state_col:
    state_df = df_f.groupby(state_col)[rev_col].sum().reset_index().rename(columns={state_col: "State", rev_col: "Revenue"})
else:
    state_df = pd.DataFrame(columns=["State", "Revenue"])

# Return Summary
if "return_status" in df_f.columns:
    df_f["is_returned"] = df_f["return_status"].astype(str).str.lower().isin(["returned", "refund", "refunded"])
    ret_df = df_f.groupby("is_returned")[rev_col].agg(["count", "sum"]).reset_index()
else:
    ret_df = pd.DataFrame(columns=["is_returned", "count", "sum"])

# -----------------------------
# ✅ FIXED: SAFE EXCEL EXPORT
# -----------------------------
def build_excel_bytes():
    import openpyxl
    output = io.BytesIO()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered_Data"

    # Always write filtered data or placeholder
    if not df_f.empty:
        ws.append(df_f.columns.tolist())
        for row in df_f.itertuples(index=False):
            ws.append(list(row))
    else:
        ws.append(["No filtered data available"])

    # KPI sheet
    ws_kpi = wb.create_sheet("KPIs")
    for i, (metric, value) in enumerate([
        ("Total Revenue", total_revenue),
        ("Total Orders", total_orders),
        ("Unique Customers", unique_customers),
        ("Avg Order Value", aov),
    ], start=1):
        ws_kpi.cell(row=i, column=1, value=metric)
        ws_kpi.cell(row=i, column=2, value=value)

    # Optional: Write other sheets only if not empty
    if not monthly.empty:
        ws_m = wb.create_sheet("Monthly_Revenue")
        ws_m.append(monthly.columns.tolist())
        for row in monthly.itertuples(index=False):
            ws_m.append(list(row))

    if not cat_df.empty:
        ws_c = wb.create_sheet("Categories")
        ws_c.append(cat_df.columns.tolist())
        for row in cat_df.itertuples(index=False):
            ws_c.append(list(row))

    if not prod_df.empty:
        ws_p = wb.create_sheet("Products")
        ws_p.append(prod_df.columns.tolist())
        for row in prod_df.itertuples(index=False):
            ws_p.append(list(row))

    if not state_df.empty:
        ws_s = wb.create_sheet("States")
        ws_s.append(state_df.columns.tolist())
        for row in state_df.itertuples(index=False):
            ws_s.append(list(row))

    if not ret_df.empty:
        ws_r = wb.create_sheet("Returns")
        ws_r.append(["is_returned", "Orders", "Revenue"])
        for i, r in ret_df.iterrows():
            ws_r.append(list(r))

    wb.save(output)
    return output.getvalue()

# Download Button
excel_data = build_excel_bytes()
st.download_button(
    "⬇️ Download Excel Report",
    data=excel_data,
    file_name="amazon_sales_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# -----------------------------
# POWERPOINT (OPTIONAL)
# -----------------------------
st.subheader("🎞 PowerPoint Export")
if not PPTX_AVAILABLE:
    st.info("Install with: `pip install python-pptx` to enable PPT export.")
else:
    st.success("✅ PowerPoint export available. (You can ask me to enable next!)")

st.success("✅ Report Export module loaded successfully.")

